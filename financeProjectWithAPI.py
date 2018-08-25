#! python3
#Author: Elena Padgett, (c) 2018

import openpyxl, os, csv, requests
from openpyxl.styles import Font

#1.a. Open and read excel worksheet from the current working directory

os.chdir('C:\\fakepath\\Python Programs')   
print (os.getcwd())              
print ('Opening workbook Data.xlsx')
wb = openpyxl.load_workbook('Data.xlsx')
sheet = wb['Sample Data']

#1.b Find and raise any exceptions: Quantity should be numbers.

def quantity(shares):
    if type(shares) != int:
        raise Exception('Quantity field can only contain digits.')

for i in range (2, sheet.max_row + 1):
    shares_value = sheet['F' + str(i)].value
    try:
        quantity(shares_value)
    except Exception as err:
        print('An exception happened: ' + str(err))

wb.save('Data_copy.xlsx')
print('Done checking for exceptions.')

#2. Calculate transaction total in local currency (Quantity * Price Local)

for i in range (2, sheet.max_row + 1):
    quantity = sheet['F' + str(i)].value
    priceLocal = sheet ['H' + str(i)].value
    totalAmount = str(quantity * priceLocal)
    sheet ['I' + str(i)] = int(totalAmount)

wb.save('Data_copy.xlsx')
print('Done calculating Total Amount in local currency.')

#3.a. Look up conversion rates using API call.

url = 'http://data.fixer.io/api/latest?access_key=xxxxxxxxxxxxxxx'  
r = requests.get(url)            
print("Status code on API call:", r.status_code)

response_dict = r.json()
print(response_dict.keys())
print("Base currency: " + response_dict['base'])

currency = response_dict['rates']   
print("Currencies:", len(currency))

for i in range(2, sheet.max_row + 1):
    curr = sheet['G' + str(i)].value
    convRate = currency[curr]
    sheet ['J' + str(i)] = convRate
   
wb.save('Data_copy.xlsx')
print('Done looking up rates.')

#3.b. Calculate Total Amount in EUR.
for i in range (2, sheet.max_row + 1): 
    totalAmountLocal = sheet['I' + str(i)].value
    convRate = sheet['J' + str(i)].value
    totalAmountEur = int(totalAmountLocal / convRate)
    sheet['K' + str(i)] = totalAmountEur

wb.save('Data_copy.xlsx')
print('Done calculating Total Amount in EUR.')

#4. Calculate Commission based on the Account number in Column A and Broker keep in basis points.

Broker_keep = {'IBM': 12, 'MCRSF': 14}

for rowNum in range(2, sheet.max_row + 1):
    accountName = sheet ['A' + str(rowNum)].value
    broker_keep = Broker_keep[accountName]
    sheet ['L' + str(rowNum)] = broker_keep 

wb.save('Data_copy.xlsx')
    
for i in range (2, sheet.max_row + 1):
    totalAmount = sheet ['K' + str(i)].value
    real_broker_keep = sheet ['L' + str(i)].value * 0.0001
    commission = totalAmount * real_broker_keep
    sheet ['M' + str(i)] = int(commission)

sheet['M12'] = '=SUM(M2:M11)'
boldTotal = Font(bold=True)
sheet['M12'].font = boldTotal

wb.save('Data_copy.xlsx')
print('Done calculating commission.')

#5. Create a csv file with how many transactions were executed on this date and how much commission earned.

totalCommission = sheet['M' + str(sheet.max_row)].value 

outputFile = open('SummaryFile.csv', 'w', newline='')
outputWriter = csv.writer(outputFile)
outputWriter.writerow(['Broker executed ' + str(sheet.max_row - 2) + ' transactions on ' + str(sheet['D2'].value) + '.'])
outputWriter.writerow(['Total Commission is $' + totalCommission])  
outputFile.close()

print('Done creating an output file Data_copy.xlsx and SummaryFile.csv.')
